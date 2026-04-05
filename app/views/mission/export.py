from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

from django.db.models import OuterRef, Subquery
from django.http import FileResponse, Http404
from django.utils import timezone
from django.views.decorators.http import require_GET
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

from app.models.document import DirectiveDocument
from app.models.mission import Department, Mission, MissionReport
from app.models.period import Period


def _parse_int_or_none(value):
    if value is None:
        return None

    value = str(value).strip()
    if not value:
        return None

    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def resolve_report_period_to_year_month(
    report_period: str | None,
) -> tuple[int | None, int | None]:
    """Kỳ từ filter: id `Period` (period_options) hoặc chuỗi `YYYY-MM` (cũ)."""
    if report_period is None:
        return None, None
    s = str(report_period).strip()
    if not s:
        return None, None
    if "-" in s:
        parts = s.split("-")
        if len(parts) == 2:
            year = _parse_int_or_none(parts[0])
            month = _parse_int_or_none(parts[1])
            if year is not None and month is not None:
                return year, month
        return None, None
    if s.isdigit():
        row = Period.objects.filter(pk=int(s)).values("year", "month").first()
        if row:
            return int(row["year"]), int(row["month"])
    return None, None


def _parse_multi_values(request, key: str) -> list[str]:
    values = request.GET.getlist(key)
    if values:
        return [str(v).strip() for v in values if str(v).strip()]

    raw = request.GET.get(key)
    if not raw:
        return []

    if isinstance(raw, str) and "," in raw:
        return [part.strip() for part in raw.split(",") if part.strip()]

    raw = str(raw).strip()
    return [raw] if raw else []


def _get_selected_report_period(request):
    report_period = (request.GET.get("report_period") or "").strip()
    y, m = resolve_report_period_to_year_month(report_period)
    if y is not None and m is not None:
        return y, m

    today = timezone.localdate()
    return today.year, today.month


def _get_report_status_label(report_status_value: str | None) -> str:
    label_map = {
        "NOT_SENT": "Chưa gửi",
        "APPROVED": "Đã gửi",
        # "NO_REPORT": "Không gửi",
    }
    return label_map.get(report_status_value, "—")


def _get_mission_status_label(mission_status_value: str | None) -> str:
    if not mission_status_value:
        return ""

    try:
        return MissionReport.MissionStatus(mission_status_value).label
    except ValueError:
        return ""


def _build_export_queryset(request):
    report_period = request.GET.get("report_period", "")
    directive_level = request.GET.get("directive_level", "")
    document_type = request.GET.get("document_type", "")
    directive_document = request.GET.get("directive_document", "")
    mission_code = (request.GET.get("name") or "").strip()
    department_values = _parse_multi_values(request, "department")
    status_values = _parse_multi_values(request, "status")
    report_status_values = _parse_multi_values(request, "report_status")

    selected_year, selected_month = resolve_report_period_to_year_month(report_period)

    report_base = MissionReport.objects.filter(mission_id=OuterRef("pk"))
    if selected_year and selected_month:
        report_base = report_base.filter(
            report_year=selected_year,
            report_month=selected_month,
        )
    else:
        report_base = report_base.order_by("-report_year", "-report_month", "-created_at")

    mission_qs = (
        Mission.objects.filter(is_active=True)
        .annotate(
            report_status=Subquery(report_base.values("status")[:1]),
            mission_status=Subquery(report_base.values("mission_status")[:1]),
            report_content=Subquery(report_base.values("content")[:1]),
        )
        .select_related(
            "department",
            "owner",
            "directive_document",
            "directive_document__directive_level",
        )
        .prefetch_related("assignee_departments")
        .distinct()
        .order_by("-created_at", "-code")
    )

    if report_period and selected_year and selected_month:
        mission_qs = mission_qs.filter(
            reports__report_year=selected_year,
            reports__report_month=selected_month,
        )

    directive_level_id = _parse_int_or_none(directive_level)
    if directive_level_id is not None:
        mission_qs = mission_qs.filter(
            directive_document__directive_level_id=directive_level_id
        )

    document_type_id = _parse_int_or_none(document_type)
    if document_type_id is not None:
        mission_qs = mission_qs.filter(directive_document__type_id=document_type_id)

    directive_document_pk = (directive_document or "").strip()
    if directive_document_pk:
        mission_qs = mission_qs.filter(directive_document_id=directive_document_pk)

    if mission_code:
        mission_qs = mission_qs.filter(code=mission_code)

    department_ids = [_parse_int_or_none(v) for v in department_values]
    department_ids = [i for i in department_ids if i is not None]
    if department_ids:
        mission_qs = mission_qs.filter(department_id__in=department_ids)

    if status_values:
        mission_qs = mission_qs.filter(mission_status__in=status_values)

    if report_status_values:
        mission_qs = mission_qs.filter(report_status__in=report_status_values)

    return mission_qs.distinct(), selected_year, selected_month


def _get_filter_display_values(request, selected_year: int, selected_month: int):
    directive_level = (request.GET.get("directive_level") or "").strip()
    directive_document_pk = (request.GET.get("directive_document") or "").strip()
    mission_code = (request.GET.get("name") or "").strip()
    department_values = _parse_multi_values(request, "department")
    status_values = _parse_multi_values(request, "status")
    report_status_values = _parse_multi_values(request, "report_status")

    directive_level_text = "Tất cả"
    directive_level_id = _parse_int_or_none(directive_level)
    if directive_level_id is not None:
        doc = (
            DirectiveDocument.objects.select_related("directive_level")
            .filter(directive_level_id=directive_level_id)
            .only("directive_level__name")
            .first()
        )
        if doc and getattr(doc, "directive_level", None):
            directive_level_text = (doc.directive_level.name or "").strip() or "Tất cả"
    directive_document_text = "Tất cả"
    if directive_document_pk:
        doc = DirectiveDocument.objects.filter(pk=directive_document_pk).first()
        if doc:
            directive_document_text = (getattr(doc, "code", "") or "").strip()

    department_text = "Tất cả"
    dept_ids = [_parse_int_or_none(v) for v in department_values]
    dept_ids = [i for i in dept_ids if i is not None]
    if dept_ids:
        names = list(
            Department.objects.filter(id__in=dept_ids)
            .order_by("short_name", "name")
            .values_list("name", flat=True)
        )
        if names:
            department_text = ", ".join(n for n in names if n)

    status_map = dict(MissionReport.MissionStatus.choices)
    status_text = ", ".join(status_map.get(v, v) for v in status_values if v) or "Tất cả"

    report_status_map = {
        "NOT_SENT": "Chưa gửi",
        "APPROVED": "Đã gửi",
        # "NO_REPORT": "Không gửi",
    }
    report_status_text = (
        ", ".join(report_status_map.get(v, v) for v in report_status_values if v) or "Tất cả"
    )

    mission_display = "Tất cả"
    if mission_code:
        m = Mission.objects.filter(pk=mission_code).only("name").first()
        mission_display = (m.name if m else mission_code).strip() or mission_code

    return {
        "directive_level": directive_level_text,
        "directive_document": directive_document_text,
        "mission_name": mission_display,
        "department": department_text,
        "status": status_text,
        "report_status": report_status_text,
        "report_period": f"{selected_month:02d}/{selected_year}",
    }


def _apply_layout(ws):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    for col in ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
        ws.column_dimensions[col].width = 10

    ws.column_dimensions["Q"].width = 14
    ws.column_dimensions["R"].width = 14
    ws.column_dimensions["S"].width = 16
    ws.column_dimensions["T"].width = 16
    ws.column_dimensions["U"].width = 26
    ws.column_dimensions["V"].width = 20

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    body_alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=22):
        for cell in row:
            if cell.row == 9:
                cell.alignment = header_alignment
            else:
                cell.alignment = body_alignment
                cell.border = thin_border

    ws.row_dimensions[9].height = 28
    for row_idx in range(10, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36


@require_GET
def mission_export_report(request):
    mission_qs, selected_year, selected_month = _build_export_queryset(request)

    if not selected_year or not selected_month:
        selected_year, selected_month = _get_selected_report_period(request)

    template_path = Path(__file__).resolve().parent / "report.xlsx"
    if not template_path.exists():
        raise Http404("Không tìm thấy file mẫu report.xlsx")

    wb = load_workbook(template_path)
    ws = wb.active

    filter_values = _get_filter_display_values(request, selected_year, selected_month)

    ws["C2"] = filter_values["report_period"]
    ws["C3"] = filter_values["mission_name"]
    ws["C4"] = filter_values["directive_level"]
    ws["C5"] = filter_values["directive_document"]
    ws["C6"] = filter_values["department"]
    ws["C7"] = filter_values["status"]
    ws["C8"] = filter_values["report_status"]

    # Header tháng: luôn đủ 12 cột của năm được chọn
    for month in range(1, 13):
        cell = ws.cell(row=10, column=4 + month)  # E -> P
        cell.value = f"{month:02d}/{selected_year}"
        cell.number_format = "@"

    # Clear data cũ từ hàng 10 trở đi
    for row_idx in range(11, max(ws.max_row, 11) + 1):
        for col_idx in range(1, 23):  # A -> V
            ws.cell(row=row_idx, column=col_idx).value = None

    mission_ids = list(mission_qs.values_list("code", flat=True))

    # Chỉ lấy báo cáo từ tháng 1 -> selected_month
    reports_by_mission_month = defaultdict(dict)
    if mission_ids:
        reports = (
            MissionReport.objects.filter(
                mission_id__in=mission_ids,
                report_year=selected_year,
                report_month__lte=selected_month,
            )
            .values("mission_id", "report_month", "content")
            .order_by("mission_id", "report_month", "-created_at", "-id")
        )

        for item in reports:
            mission_id = item["mission_id"]
            month = int(item["report_month"])
            if month not in reports_by_mission_month[mission_id]:
                reports_by_mission_month[mission_id][month] = item.get("content") or ""

    period_text = f"{selected_month:02d}/{selected_year}"
    start_row = 11

    for index, mission in enumerate(mission_qs, start=1):
        row_idx = start_row + index - 1

        assignee_names = ", ".join(
            d.get_short_label()
            for d in mission.assignee_departments.all()
            if d.get_short_label()
        ) or "—"

        ws.cell(row=row_idx, column=1, value=index)
        ws.cell(row=row_idx, column=2, value=period_text)
        ws.cell(
            row=row_idx,
            column=3,
            value=(getattr(mission.directive_document, "code", "") or "").strip()
            if getattr(mission, "directive_document", None)
            else "",
        )
        ws.cell(row=row_idx, column=4, value=mission.name or "")

        # Chỉ đổ dữ liệu tới tháng được chọn, các tháng sau để trống
        for month in range(1, 13):
            value = ""
            if month <= selected_month:
                value = reports_by_mission_month.get(mission.code, {}).get(month, "")

            ws.cell(
                row=row_idx,
                column=4 + month,
                value=value,
            )

        due_cell = ws.cell(row=row_idx, column=17, value=getattr(mission, "due_date", None))
        done_cell = ws.cell(row=row_idx, column=18, value=getattr(mission, "completed_date", None))
        due_cell.number_format = "dd/mm/yyyy"
        done_cell.number_format = "dd/mm/yyyy"

        ws.cell(
            row=row_idx,
            column=19,
            value=_get_mission_status_label(getattr(mission, "mission_status", None)),
        )
        ws.cell(
            row=row_idx,
            column=20,
            value=mission.department.get_short_label() if getattr(mission, "department", None) else "",
        )
        ws.cell(row=row_idx, column=21, value=assignee_names)
        ws.cell(
            row=row_idx,
            column=22,
            value=_get_report_status_label(getattr(mission, "report_status", None)),
        )

    _apply_layout(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Bao_cao_ket_qua_thuc_hien_nhiem_vu_{selected_year}_{selected_month:02d}.xlsx"
    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )