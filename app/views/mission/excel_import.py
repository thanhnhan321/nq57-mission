from datetime import datetime, date
from html import escape as html_escape
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.db import IntegrityError, transaction
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.utils import timezone

from app.models import Department, Mission
from app.models.document import DirectiveDocument
from app.models.mission import MissionReport
from app.views.templates.components.table import TableColumn

from .create import MissionCreateView


@require_GET
def mission_template_download(request):
    template_path = Path(__file__).resolve().parent / "Mission.xlsx"
    if not template_path.exists():
        raise Http404("Template file not found.")

    # Đọc file mẫu có sẵn
    wb = load_workbook(template_path)

    # Lấy sheet hướng dẫn
    if "hướng dẫn" not in wb.sheetnames:
        raise Http404("Template sheet 'hướng dẫn' not found.")

    ws = wb["hướng dẫn"]

    # Chỉ lấy các đơn vị thuộc CAP, CAX
    departments = Department.objects.filter(type__in=["CAP", "CAX"]).order_by("type", "id")

    # Bảng đơn vị:
    # Header ở hàng 4, dữ liệu bắt đầu từ hàng 5
    # Cột H -> K tương ứng: STT | Tên đơn vị | Tên viết tắt | Nhóm đơn vị
    start_row = 5
    start_col = 8  # H

    # Xóa dữ liệu cũ vùng H5:K...
    max_clear_row = max(ws.max_row, start_row + departments.count() + 50)
    for row in ws.iter_rows(min_row=start_row, max_row=max_clear_row, min_col=8, max_col=11):
        for cell in row:
            cell.value = None
            cell.border = Border()
            cell.alignment = Alignment()

    # Border mảnh
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Alignment
    center_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(vertical="center", wrap_text=True)
    normal_alignment = Alignment(vertical="center")

    # Set border + căn giữa cho header H4:K4
    for col in range(8, 12):
        header_cell = ws.cell(row=4, column=col)
        header_cell.border = border
        header_cell.alignment = center_alignment

    # Ghi dữ liệu đơn vị
    for idx, dept in enumerate(departments, start=1):
        row = start_row + idx - 1

        dept_name = getattr(dept, "name", "") or ""
        dept_short_name = getattr(dept, "short_name", "") or ""
        group_value = (getattr(dept, "type", "") or "").strip().upper()

        stt_cell = ws.cell(row=row, column=8, value=idx)                # H
        name_cell = ws.cell(row=row, column=9, value=dept_name)         # I
        short_cell = ws.cell(row=row, column=10, value=dept_short_name) # J
        group_cell = ws.cell(row=row, column=11, value=group_value)     # K

        # Border
        for cell in (stt_cell, name_cell, short_cell, group_cell):
            cell.border = border

        # Alignment
        stt_cell.alignment = center_alignment
        name_cell.alignment = wrap_alignment
        short_cell.alignment = normal_alignment
        group_cell.alignment = center_alignment

    # Set width cột đẹp hơn
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 80
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 14

    # Tăng chiều cao dòng cho các dòng có thể wrap text
    last_data_row = start_row + departments.count() - 1
    if last_data_row >= start_row:
        for row in range(start_row, last_data_row + 1):
            ws.row_dimensions[row].height = 35

    # Xuất file sau xử lý
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Mission.xlsx"'
    return response


@csrf_exempt
@require_POST
def mission_excel_validate(request):
    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        return HttpResponse(
            "<div class='mt-3 rounded-lg border border-red-200 bg-red-50 p-3 text-sm text-red-700'>"
            "Chưa có file Excel để kiểm tra."
            "</div>"
        )

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.datetime import from_excel
    except ImportError:
        return HttpResponse(
            "<div class='mt-3 rounded-lg border border-red-200 bg-red-50 p-3 text-sm text-red-700'>"
            "Server chưa cài `openpyxl` để đọc file Excel."
            "</div>"
        )

    try:
        in_memory = excel_file.read()
        wb = load_workbook(filename=BytesIO(in_memory), data_only=True)
        ws = wb.active
    except Exception:
        return HttpResponse(
            "<div class='mt-3 rounded-lg border border-red-200 bg-red-50 p-3 text-sm text-red-700'>"
            "File Excel không hợp lệ hoặc không đọc được."
            "</div>"
        )

    cap_short_names = set(
        s.upper()
        for s in Department.objects.filter(type=Department.Type.CAP)
        .values_list("short_name", flat=True)
        .iterator()
        if s
    )
    all_short_names = set(
        s.upper()
        for s in Department.objects.filter(short_name__isnull=False)
        .values_list("short_name", flat=True)
        .iterator()
        if s
    )

    first_day_of_current_month = timezone.localdate().replace(day=1)

    def _to_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        return str(v).strip()

    def _parse_ddmmyyyy(v):
        if v is None or _to_str(v) == "":
            return None

        if hasattr(v, "date"):
            try:
                return v.date()
            except Exception:
                pass

        if isinstance(v, (int, float)):
            try:
                return from_excel(v).date()
            except Exception:
                return None

        s = _to_str(v)
        try:
            return datetime.strptime(s, "%d/%m/%Y").date()
        except ValueError:
            return None

    def _split_short_names(v):
        s = _to_str(v)
        if not s:
            return []
        return [part.strip() for part in s.split(",") if part.strip()]

    def _format_date_display(date_value, raw_text: str) -> str:
        if date_value is None:
            return raw_text
        try:
            return date_value.strftime("%d/%m/%Y")
        except Exception:
            return raw_text

    max_row = getattr(ws, "max_row", 0) or 0
    stt = 1
    error_rows: list[dict] = []

    for r in range(2, max_row + 1):
        a_raw = ws.cell(r, 1).value
        b_raw = ws.cell(r, 2).value
        c_raw = ws.cell(r, 3).value
        d_raw = ws.cell(r, 4).value
        e_raw = ws.cell(r, 5).value

        if all(_to_str(x) == "" for x in (a_raw, b_raw, c_raw, d_raw, e_raw)):
            break

        a = _to_str(a_raw)
        b = _to_str(b_raw)
        c = _to_str(c_raw)
        d = _to_str(d_raw)

        c_date = _parse_ddmmyyyy(c_raw)
        d_date = _parse_ddmmyyyy(d_raw)
        e_parts = _split_short_names(e_raw)

        row_errors: list[str] = []

        if not a:
            row_errors.append("Cột A không được rỗng")

        if not b:
            row_errors.append("Cột B Mã ĐV Chủ trì không được rỗng")
        elif b.upper() not in cap_short_names:
            row_errors.append("Cột B đơn vị chủ trì phải thuộc CAP")

        if not c:
            row_errors.append("Cột C ngày bắt đầu không được rỗng")
        elif c_date is None:
            row_errors.append("Cột C định dạng ngày phải là dd/mm/yyyy")
        elif c_date < first_day_of_current_month:
            row_errors.append(
                f"Cột C ngày bắt đầu không được nhỏ hơn ngày {first_day_of_current_month.strftime('%d/%m/%Y')}"
            )

        if d:
            if d_date is None:
                row_errors.append("Cột D định dạng ngày phải là dd/mm/yyyy")
            elif c_date is not None and d_date <= c_date:
                row_errors.append("Cột D ngày kết thúc phải lớn hơn ngày bắt đầu")

        if not e_parts:
            row_errors.append("Cột E Đơn vị thực hiện không được rỗng")
        else:
            missing = [sn for sn in e_parts if sn.upper() not in all_short_names]
            if missing:
                row_errors.append(f"Thiếu tên viết tắt đơn vị ở Cột E: {', '.join(missing)}")

        error_rows.append(
            {
                "stt": stt,
                "col_a": a,
                "col_b": b,
                "col_c": _format_date_display(c_date, c),
                "col_d": _format_date_display(d_date, d),
                "col_e": ", ".join(e_parts),
                "errors": (
                    f"<span class='text-red-700 font-medium'>{html_escape(' ; '.join(row_errors))}</span>"
                    if row_errors
                    else ""
                ),
            }
        )
        stt += 1

    has_errors = any((row.get("errors") or "").strip() for row in error_rows)
    request.session["mission_excel_validation_errors"] = error_rows

    page_index = 0
    page_size = 10
    total_count = len(error_rows)
    rows_page = error_rows[:page_size]

    columns = [
        TableColumn(name="stt", label="#", sortable=False),
        TableColumn(name="col_a", label="Cột A", sortable=False),
        TableColumn(name="col_b", label="Cột B", sortable=False),
        TableColumn(name="col_c", label="Cột C", sortable=False),
        TableColumn(name="col_d", label="Cột D", sortable=False),
        TableColumn(name="col_e", label="Cột E", sortable=False),
        TableColumn(name="errors", label="Ghi chú lỗi", sortable=False),
    ]

    context = {
        "columns": columns,
        "rows": rows_page,
        "sort": "",
        "sort_direction": "asc",
        "partial_url": reverse("mission_excel_validate_page"),
        "actions": [],
        "has_errors": has_errors,
        "bulk_actions": [],
        "row_actions": [],
        "page_index": page_index,
        "page_size": page_size,
        "total_count": total_count,
    }
    return render(request, "mission/mission_excel_validation_table.html", context)


@require_GET
def mission_excel_validate_page(request):
    all_rows: list[dict] = request.session.get("mission_excel_validation_errors", []) or []

    try:
        page_index = int(request.GET.get("page_index", "0") or 0)
    except ValueError:
        page_index = 0

    try:
        page_size = int(request.GET.get("page_size", "10") or 10)
    except ValueError:
        page_size = 10

    page_index = max(0, page_index)
    page_size = max(1, page_size)

    start = page_index * page_size
    end = start + page_size
    rows_page = all_rows[start:end]

    columns = [
        TableColumn(name="stt", label="#", sortable=False),
        TableColumn(name="col_a", label="Cột A", sortable=False),
        TableColumn(name="col_b", label="Cột B", sortable=False),
        TableColumn(name="col_c", label="Cột C", sortable=False),
        TableColumn(name="col_d", label="Cột D", sortable=False),
        TableColumn(name="col_e", label="Cột E", sortable=False),
        TableColumn(name="errors", label="Ghi chú lỗi", sortable=False),
    ]

    context = {
        "columns": columns,
        "rows": rows_page,
        "sort": "",
        "sort_direction": "asc",
        "partial_url": reverse("mission_excel_validate_page"),
        "bulk_actions": [],
        "row_actions": [],
        "page_index": page_index,
        "page_size": page_size,
        "total_count": len(all_rows),
    }
    return render(request, "mission/mission_excel_validation_partial.html", context)


@csrf_exempt
@require_POST
def mission_excel_create(request):
    """
    Create `Mission` + `MissionReport` from the validated Excel rows stored in session.

    Excel column mapping:
    - Cột A (col_a): Mission name
    - Cột B (col_b): CAP department short_name (Mission.department)
    - Cột C (col_c): start_date (dd/mm/YYYY)
    - Cột D (col_d): due_date (dd/mm/YYYY, optional)
    - Cột E (col_e): assignee_departments short_names list (comma separated)
    """

    directive_level_id = (request.POST.get("directive_level") or "").strip()
    directive_document_pk = (request.POST.get("directive_document") or "").strip()

    try:
        directive_level_id_int = int(directive_level_id)
    except (TypeError, ValueError):
        return JsonResponse(
            {"success": False, "message": "Cấp chỉ đạo không hợp lệ."},
            status=400,
        )

    if not directive_document_pk:
        return JsonResponse(
            {"success": False, "message": "Văn bản chỉ đạo là bắt buộc."},
            status=400,
        )

    error_rows: list[dict] = request.session.get("mission_excel_validation_errors", []) or []
    if not error_rows:
        return JsonResponse(
            {
                "success": False,
                "message": "Chưa có dữ liệu Excel để tạo. Vui lòng Upload & kiểm tra trước.",
            },
            status=400,
        )

    has_errors = any((row.get("errors") or "").strip() for row in error_rows)
    if has_errors:
        messages.error(request, "File Excel còn lỗi, vui lòng sửa trước khi xác nhận.")
        return JsonResponse(
            {"success": False, "message": "File Excel còn lỗi, vui lòng sửa trước khi xác nhận."},
            status=400,
        )

    def _parse_ddmmyyyy(s: str | None) -> date | None:
        s = (s or "").strip()
        if not s:
            return None
        try:
            return datetime.strptime(s, "%d/%m/%Y").date()
        except ValueError:
            return None

    doc = (
        DirectiveDocument.objects.filter(pk=directive_document_pk)
        .select_related("directive_level")
        .first()
    )
    if not doc:
        return JsonResponse(
            {"success": False, "message": "Văn bản chỉ đạo không tồn tại."},
            status=400,
        )

    if doc.directive_level_id != directive_level_id_int:
        return JsonResponse(
            {
                "success": False,
                "message": "Văn bản chỉ đạo không thuộc cấp chỉ đạo đã chọn.",
            },
            status=400,
        )

    departments = Department.objects.filter(short_name__isnull=False).all()
    dept_by_short_upper = {
        (d.short_name or "").strip().upper(): d
        for d in departments
        if (d.short_name or "").strip()
    }

    creator = MissionCreateView()
    created_count = 0

    try:
        with transaction.atomic():
            for row in error_rows:
                owner_short = (row.get("col_b") or "").strip().upper()
                if not owner_short:
                    raise ValueError("Thiếu Cột B (đơn vị chủ trì) trong dữ liệu Excel.")

                owner_dept = dept_by_short_upper.get(owner_short)
                if not owner_dept:
                    raise ValueError(
                        f"Không tìm thấy đơn vị chủ trì với short_name: {owner_short}"
                    )

                assignee_shorts = [
                    sn.strip().upper()
                    for sn in str(row.get("col_e") or "").split(",")
                    if sn.strip()
                ]

                seen: set[str] = set()
                assignee_shorts = [
                    sn for sn in assignee_shorts if not (sn in seen or seen.add(sn))
                ]

                assignee_depts = []
                for sn in assignee_shorts:
                    d = dept_by_short_upper.get(sn)
                    if not d:
                        raise ValueError(
                            f"Không tìm thấy đơn vị thực hiện với short_name: {sn}"
                        )
                    assignee_depts.append(d)

                name = (row.get("col_a") or "").strip()
                if not name:
                    raise ValueError("Thiếu Cột A (tên nhiệm vụ) trong dữ liệu Excel.")

                start_date = _parse_ddmmyyyy(row.get("col_c"))
                if not start_date:
                    raise ValueError("Cột C (ngày bắt đầu) không đúng định dạng.")
                first_day_of_current_month = timezone.localdate().replace(day=1)
                if start_date < first_day_of_current_month:
                    raise ValueError(
                        f"Ngày bắt đầu không được nhỏ hơn ngày {first_day_of_current_month.strftime('%d/%m/%Y')}."
                    )

                due_date = (
                    _parse_ddmmyyyy(row.get("col_d"))
                    if (row.get("col_d") or "").strip()
                    else None
                )
                if due_date and due_date < start_date:
                    raise ValueError("Hạn xử lý không được nhỏ hơn ngày bắt đầu.")

                mission = None
                last_err: Exception | None = None

                for _ in range(5):
                    try:
                        mission = Mission(
                            code=creator._generate_mission_code(department=owner_dept),
                            name=name,
                            start_date=start_date,
                            due_date=due_date,
                            department=owner_dept,
                            directive_document_id=directive_document_pk,
                            created_by_id=request.user.id,
                            updated_by_id=request.user.id,
                        )
                        mission.save()
                        last_err = None
                        break
                    except IntegrityError as e:
                        last_err = e

                if mission is None or last_err is not None:
                    raise ValueError("Không thể sinh mã nhiệm vụ. Vui lòng thử lại.")

                mission.assignee_departments.set(assignee_depts)

                report_year = int(start_date.year)
                report_month = int(start_date.month)
                if int(start_date.day) < 10:
                    if report_month == 1:
                        report_month = 12
                        report_year -= 1
                    else:
                        report_month -= 1

                MissionReport.objects.get_or_create(
                    mission=mission,
                    report_year=report_year,
                    report_month=report_month,
                    defaults={},
                )

                created_count += 1

    except ValueError as e:
        messages.error(request, str(e))
        return JsonResponse({"success": False, "message": str(e)}, status=400)
    except Exception as e:
        messages.error(request, f"Lỗi: {str(e)}")
        return JsonResponse(
            {"success": False, "message": f"Lỗi: {str(e)}"},
            status=500,
        )

    request.session.pop("mission_excel_validation_errors", None)
    messages.success(request, f"Tạo {created_count} nhiệm vụ từ Excel thành công.")
    return JsonResponse({"success": True, "created_count": created_count})