from __future__ import annotations

from collections import defaultdict
from copy import copy
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

from app.models.department import Department
from app.models.department_report import DepartmentReport
from app.models.period import Period
from .filters import parse_int_or_none


STATUS_LABEL_MAP = {
    "SENT": "Đã gửi",
    "NO_REPORT": "Không gửi",
}

REPORT_TYPE_LABEL_MAP = {
    "MONTH": "Báo cáo tháng",
    "QUARTER": "Báo cáo quý",
    "HALF_YEAR": "Báo cáo 6 tháng",
    "NINE_MONTH": "Báo cáo 9 tháng",
    "YEAR": "Báo cáo năm",
}

THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def _is_report_admin(user) -> bool:
    return bool(getattr(user, "is_superuser", False))


def _get_user_department_id(user):
    profile = getattr(user, "profile", None)
    department_id = getattr(profile, "department_id", None)
    if department_id:
        return department_id
    return None


def _format_period(month: int | None, year: int | None) -> str:
    if month and year:
        return f"{int(month):02d}/{int(year)}"
    if year:
        return str(year)
    return ""


def _status_label(value: str | None) -> str:
    return STATUS_LABEL_MAP.get(value, value or "")


def _report_type_label(value: str | None) -> str:
    return REPORT_TYPE_LABEL_MAP.get(value, value or "")


def _department_full_name(department) -> str:
    if not department:
        return ""
    return getattr(department, "name", str(department))


def _excel_safe_datetime(value):
    if not value:
        return value
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.replace(tzinfo=None)


def _copy_cell_style(ws, from_row: int, to_row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        source = ws.cell(from_row, col)
        target = ws.cell(to_row, col)

        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)

        if source._style:
            target._style = copy(source._style)


def _prepare_rows(ws, start_row: int, total_rows: int, start_col: int, end_col: int) -> None:
    if total_rows <= 1:
        return

    ws.insert_rows(start_row + 1, total_rows - 1)

    for row in range(start_row + 1, start_row + total_rows):
        _copy_cell_style(ws, start_row, row, start_col, end_col)


def _normalize_period_ids(request) -> list[int]:
    raw_values = request.GET.getlist("period_ids")
    result: list[int] = []

    if not raw_values:
        raw_value = str(request.GET.get("period_ids", "")).strip()
        if raw_value:
            raw_values = [part.strip() for part in raw_value.split(",") if part.strip()]

    for value in raw_values:
        value = str(value or "").strip()
        if value.isdigit():
            result.append(int(value))

    unique_result = []
    seen = set()
    for item in result:
        if item not in seen:
            seen.add(item)
            unique_result.append(item)

    return unique_result


def _normalize_status_values(request) -> list[str]:
    raw_status_values = request.GET.getlist("status")
    if raw_status_values:
        return [str(v).strip() for v in raw_status_values if str(v).strip()]

    raw_status = request.GET.get("status", "")
    if isinstance(raw_status, str) and "," in raw_status:
        return [part.strip() for part in raw_status.split(",") if part.strip()]

    raw_status = str(raw_status).strip()
    return [raw_status] if raw_status else []


def _get_period_objects(period_ids: list[int]) -> list[Period]:
    if not period_ids:
        return []

    period_map = {obj.id: obj for obj in Period.objects.filter(id__in=period_ids)}
    return [period_map[pid] for pid in period_ids if pid in period_map]


def _period_label_from_obj(period: Period) -> str:
    if getattr(period, "name", None):
        return str(period.name)
    if getattr(period, "code", None):
        return str(period.code)

    month = getattr(period, "month", None)
    year = getattr(period, "year", None)
    if month and year:
        return f"Tháng {int(month):02d}/{int(year)}"
    if year:
        return str(year)

    return str(period.pk)


def _get_period_label(period_ids: list[int]) -> str:
    periods = _get_period_objects(period_ids)
    if not periods:
        return "Tất cả"

    labels = [_period_label_from_obj(period) for period in periods]
    return ", ".join([label for label in labels if label]) or "Tất cả"


def _get_department_label(department_id: int | None) -> str:
    if department_id is None:
        return "Tất cả"

    obj = Department.objects.filter(id=department_id).first()
    if obj:
        return _department_full_name(obj)

    return "Tất cả"


def _join_status_labels(status_values: list[str]) -> str:
    if not status_values:
        return "Đã gửi"
    return ", ".join(_status_label(v) for v in status_values)


def _build_queryset(request):
    period_ids = _normalize_period_ids(request)
    report_type = request.GET.get("report_type", "")
    status_values = _normalize_status_values(request)

    is_admin = _is_report_admin(request.user)
    user_department_id = _get_user_department_id(request.user)

    if is_admin:
        department = request.GET.get("department", "")
        department_id = parse_int_or_none(department)
    else:
        department_id = user_department_id

    qs = (
        DepartmentReport.objects
        .select_related("department", "file", "period")
        .order_by("department__name", "department_id", "report_type", "id")
    )

    if period_ids:
        qs = qs.filter(period_id__in=period_ids)

    if department_id is not None:
        qs = qs.filter(department_id=department_id)

    if status_values:
        qs = qs.filter(status__in=status_values)

    if report_type:
        qs = qs.filter(report_type=report_type)

    return qs, period_ids, department_id, status_values, report_type


def _period_label_from_report(obj: DepartmentReport) -> str:
    if getattr(obj, "period", None):
        return _period_label_from_obj(obj.period)
    return _format_period(obj.month, obj.report_year)


def _safe_merge_row(ws, row: int, start_col: int, end_col: int) -> None:
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    target_range = f"{start_letter}{row}:{end_letter}{row}"

    merged_to_remove = []
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        if min_row <= row <= max_row and not (end_col < min_col or start_col > max_col):
            merged_to_remove.append(str(merged))

    for merged_range in merged_to_remove:
        ws.unmerge_cells(merged_range)

    ws.merge_cells(target_range)


def _style_filter_value_cell(ws, cell_ref: str) -> None:
    cell = ws[cell_ref]
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.font = copy(cell.font) if cell.font else Font(name="Times New Roman", size=12)


def _fill_filter_cells(
    ws,
    period_label: str,
    department_label: str,
    report_type_label: str,
    status_label: str,
) -> None:
    _safe_merge_row(ws, 2, 3, 8)  # C2:H2
    _safe_merge_row(ws, 3, 3, 8)  # C3:H3
    _safe_merge_row(ws, 4, 3, 8)  # C4:H4
    _safe_merge_row(ws, 5, 3, 8)  # C5:H5

    ws["C2"] = period_label
    ws["C3"] = department_label
    ws["C4"] = report_type_label
    ws["C5"] = status_label

    _style_filter_value_cell(ws, "C2")
    _style_filter_value_cell(ws, "C3")
    _style_filter_value_cell(ws, "C4")
    _style_filter_value_cell(ws, "C5")

    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24


def _set_detail_layout(ws) -> None:
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14


def _set_summary_layout(ws) -> None:
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14


def _apply_table_borders(ws, start_row: int, total_rows: int, start_col: int, end_col: int) -> None:
    if total_rows <= 0:
        total_rows = 1

    for row in range(start_row, start_row + total_rows):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = THIN_BORDER


def _apply_center_alignment(ws, row: int, cols: list[int]) -> None:
    for col in cols:
        ws.cell(row, col).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _apply_left_alignment(ws, row: int, cols: list[int]) -> None:
    for col in cols:
        ws.cell(row, col).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )


def _write_detail_section(ws, rows: list[DepartmentReport], start_row: int = 9) -> None:
    for index, obj in enumerate(rows, start=1):
        row = start_row + index - 1

        ws.cell(row, 1).value = index
        ws.cell(row, 2).value = _period_label_from_report(obj)
        ws.cell(row, 3).value = _report_type_label(obj.report_type)
        ws.cell(row, 4).value = _department_full_name(obj.department)

        if obj.sent_at:
            ws.cell(row, 5).value = _excel_safe_datetime(obj.sent_at)
            ws.cell(row, 5).number_format = "dd/mm/yyyy"
        else:
            ws.cell(row, 5).value = ""

        ws.cell(row, 6).value = _status_label(obj.status)

        _apply_center_alignment(ws, row, [1, 2, 5, 6])
        _apply_left_alignment(ws, row, [3, 4])


def _build_filename(prefix: str, period_label: str) -> str:
    if not period_label or period_label.strip().lower() == "tất cả":
        return f"{prefix}.xlsx"

    safe_period = (
        period_label
        .replace("/", "_")
        .replace(", ", "__")
        .strip()
    )
    return f"{prefix}_{safe_period}.xlsx"


def _build_summary_rows(rows: list[DepartmentReport]) -> list[dict]:
    """
    Logic mới:
    - Mỗi dòng = 1 nhóm (kỳ báo cáo, loại báo cáo)
    - Đã gửi = số đơn vị distinct có status = SENT
    - Không gửi = số đơn vị distinct có status = NO_REPORT
    - Query trực tiếp từ bảng theo status, không lấy tổng đơn vị - đã gửi nữa
    """
    grouped_sent_departments: dict[tuple[str, str], set[int]] = defaultdict(set)
    grouped_no_report_departments: dict[tuple[str, str], set[int]] = defaultdict(set)
    grouped_all_seen: set[tuple[str, str]] = set()

    for obj in rows:
        key = (
            _period_label_from_report(obj),
            _report_type_label(obj.report_type),
        )
        grouped_all_seen.add(key)

        if obj.department_id:
            if obj.status == "SENT":
                grouped_sent_departments[key].add(obj.department_id)
            elif obj.status == "NO_REPORT":
                grouped_no_report_departments[key].add(obj.department_id)

    result = []
    for period_label, report_type_label in sorted(
        grouped_all_seen,
        key=lambda item: (item[0], item[1]),
    ):
        key = (period_label, report_type_label)
        sent_count = len(grouped_sent_departments.get(key, set()))
        not_sent_count = len(grouped_no_report_departments.get(key, set()))

        result.append(
            {
                "period_label": period_label,
                "report_type_label": report_type_label,
                "sent": sent_count,
                "not_sent": not_sent_count,
            }
        )

    return result


def _write_summary_section(ws, rows: list[dict], start_row: int = 9) -> None:
    for index, item in enumerate(rows, start=1):
        row = start_row + index - 1

        ws.cell(row, 1).value = index
        ws.cell(row, 2).value = item["period_label"]
        ws.cell(row, 3).value = item["report_type_label"]
        ws.cell(row, 4).value = item["sent"]
        ws.cell(row, 5).value = item["not_sent"]

        _apply_center_alignment(ws, row, [1, 2, 4, 5])
        _apply_left_alignment(ws, row, [3])


def export_detail_report_excel(request):
    template_path = Path(__file__).resolve().parent / "detail_report.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file mẫu: {template_path}")

    qs, period_ids, department_id, status_values, report_type = _build_queryset(request)
    rows = list(qs)

    wb = load_workbook(template_path)
    ws = wb.active

    period_label = _get_period_label(period_ids)
    department_label = _get_department_label(department_id)
    report_type_label = "Tất cả" if not report_type else _report_type_label(report_type)
    status_label = _join_status_labels(status_values)

    _fill_filter_cells(
        ws=ws,
        period_label=period_label,
        department_label=department_label,
        report_type_label=report_type_label,
        status_label=status_label,
    )
    _set_detail_layout(ws)

    data_start_row = 8
    total_rows = max(len(rows), 1)

    _prepare_rows(ws, start_row=data_start_row, total_rows=total_rows, start_col=1, end_col=6)
    _write_detail_section(ws, rows, start_row=data_start_row)
    _apply_table_borders(ws, start_row=data_start_row, total_rows=total_rows, start_col=1, end_col=6)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = _build_filename("bao_cao_chi_tiet", period_label)
    quoted_filename = quote(filename)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=export.xlsx; filename*=UTF-8''{quoted_filename}"
    )
    return response


def export_summary_report_excel(request):
    template_path = Path(__file__).resolve().parent / "summary_report.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file mẫu: {template_path}")

    # Summary giờ lấy trực tiếp theo queryset đã filter
    qs, period_ids, department_id, status_values, report_type = _build_queryset(request)
    rows = list(
        qs.select_related("department", "file", "period").order_by(
            "department__name", "department_id", "report_type", "id"
        )
    )

    summary_rows = _build_summary_rows(rows)

    wb = load_workbook(template_path)
    ws = wb.active

    period_label = _get_period_label(period_ids)
    department_label = _get_department_label(department_id)
    report_type_label = "Tất cả" if not report_type else _report_type_label(report_type)
    status_label = _join_status_labels(status_values)

    _fill_filter_cells(
        ws=ws,
        period_label=period_label,
        department_label=department_label,
        report_type_label=report_type_label,
        status_label=status_label,
    )
    _set_summary_layout(ws)

    data_start_row = 9
    total_rows = max(len(summary_rows), 1)

    _prepare_rows(ws, start_row=data_start_row, total_rows=total_rows, start_col=1, end_col=5)
    _write_summary_section(ws, summary_rows, start_row=data_start_row)
    _apply_table_borders(ws, start_row=data_start_row, total_rows=total_rows, start_col=1, end_col=5)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = _build_filename("bao_cao_tong_hop", period_label)
    quoted_filename = quote(filename)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=export.xlsx; filename*=UTF-8''{quoted_filename}"
    )
    return response