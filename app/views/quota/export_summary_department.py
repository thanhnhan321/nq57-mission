from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from django.db.models import Count, F, Q, Sum
from django.db.models.base import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

from ...models import Period, Quota, QuotaAssignment, QuotaReport


def _parse_period_ids(request) -> list[int]:
    raw = [v for v in request.GET.getlist("period_ids", []) if v]
    ids = [int(x) for x in raw if str(x).isdigit()]
    if ids:
        return ids

    today = timezone.localdate()
    p = Period.objects.filter(year=today.year, month=today.month).values_list("id", flat=True).first()
    return [p] if p is not None else []


def _period_label(period: Period) -> str:
    return f"Tháng {period.month:02d}/{period.year}"


class ExportSummaryDepartmentView(View):
    """
    Export Excel tổng hợp theo "đơn vị chủ trì".

    Logic đánh giá theo mô tả:
    - Với mỗi `dv chủ trì` và mỗi tháng, xét các chỉ tiêu (quota) mà đơn vị đó chủ trì.
    - Đánh giá "Đạt/Không đạt" theo từng chỉ tiêu trong tháng đó (tính bằng tổng expected/actual của
      tất cả `đơn vị thực hiện` đã có `QuotaReport` cho quota đó trong tháng).
    - `Số lượng đơn vị thực hiện tối đa` = số lượng `đơn vị thực hiện` distinct xuất hiện trong các chỉ tiêu
      (quota) đó của tháng.
    """

    TEMPLATE_PATH = Path(__file__).resolve().parent / "summary_department_quota.xlsx"

    def _get_filter_values(self, request) -> dict[str, Any]:
        filters = request.GET
        period_ids = _parse_period_ids(request)
        period_label = "Tất cả"
        if period_ids:
            periods = list(Period.objects.filter(id__in=period_ids).order_by("-year", "-month"))
            period_label = ", ".join(_period_label(p) for p in periods)

        return {
            "name": filters.get("name", "Tất cả"),
            "lead_department": filters.get("lead_department", "Tất cả"),
            "assigned_department": filters.get("assigned_department", "Tất cả"),
            "evaluation_result": filters.get("evaluation_result", "Tất cả"),
            "report_statuses": ", ".join(filters.getlist("report_statuses")) or "Tất cả",
            "period_label": period_label,
        }

    def get(self, request, *args, **kwargs):
        if not self.TEMPLATE_PATH.exists():
            return HttpResponse("Không tìm thấy file mẫu Excel.", status=404)

        period_ids = _parse_period_ids(request)
        periods = list(Period.objects.filter(id__in=period_ids).order_by("-year", "-month"))
        period_label_by_id = {p.id: _period_label(p) for p in periods}

        context_filters = self._get_filter_values(request)

        # =====================
        # 1. XỬ LÝ QUERY
        # =====================
        name_text = (request.GET.get("name") or "").strip()
        lead_department_id = (request.GET.get("lead_department") or "").strip()
        evaluation_result_raw = (request.GET.get("evaluation_result") or "").strip()
        report_statuses = [v for v in request.GET.getlist("report_statuses") if v]

        evaluation_filter: bool | None = None
        if evaluation_result_raw in {"true", "false"}:
            evaluation_filter = evaluation_result_raw == "true"

        quota_qs = Quota.objects.all()

        # Luôn giới hạn theo các tháng đã chọn để tránh kéo dữ liệu không liên quan.
        if period_ids:
            quota_qs = quota_qs.filter(department_reports__period_id__in=period_ids)

        # Quyền truy cập tương tự logic ở `get_common_context` (quota list).
        if not request.user.is_superuser:
            dept_id = request.user.profile.department_id
            quota_qs = quota_qs.filter(
                Q(department_assignments__department_id=dept_id, department_assignments__is_leader=True)
                | Q(
                    department_reports__department_id=dept_id,
                    department_reports__period_id__in=period_ids,
                )
            ).distinct()

        if name_text:
            quota_qs = quota_qs.filter(name__icontains=name_text)

        if lead_department_id.isdigit():
            quota_qs = quota_qs.filter(
                department_assignments__department_id=lead_department_id,
                department_assignments__is_leader=True,
            )

        if report_statuses:
            quota_qs = quota_qs.filter(
                department_reports__period_id__in=period_ids,
                department_reports__status__in=report_statuses,
            ).distinct()

        quota_ids = list(quota_qs.values_list("id", flat=True))
        if not quota_ids:
            # Vẫn xuất file để người dùng tải xuống.
            wb = load_workbook(self.TEMPLATE_PATH)
            ws = wb.active

            ws["C2"] = context_filters["name"]
            ws["C3"] = context_filters["lead_department"]
            ws["C4"] = context_filters["assigned_department"]
            ws["C5"] = context_filters["evaluation_result"]
            ws["C6"] = context_filters["report_statuses"]
            ws["C7"] = context_filters["period_label"]

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            vn_filename = "Báo cáo tổng hợp kết quả thực hiện chỉ tiêu theo đơn vị chủ trì.xlsx"
            ascii_filename = "Bao_cao_tong_hop_ket_qua_thuc_hien_chi_tieu_theo_don_vi_chu_tri.xlsx"
            encoded_vn_filename = quote(vn_filename)
            return HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": (
                        "attachment;"
                        f' filename="{ascii_filename}";'
                        f" filename*=UTF-8''{encoded_vn_filename}"
                    )
                },
            )

        leader_by_quota: dict[int, dict[str, Any]] = {}
        leader_rows = (
            QuotaAssignment.objects.filter(is_leader=True, quota_id__in=quota_ids)
            .values("quota_id", "department_id", "department__short_name")
            .distinct()
        )
        for r in leader_rows:
            # Nếu có dữ liệu bất thường nhiều hơn 1 dv chủ trì cho cùng quota,
            # lấy cái đầu tiên để tránh đếm trùng.
            leader_by_quota.setdefault(
                int(r["quota_id"]),
                {
                    "lead_department_id": int(r["department_id"]),
                    "lead_department_name": r["department__short_name"] or "",
                },
            )

        if not leader_by_quota:
            wb = load_workbook(self.TEMPLATE_PATH)
            ws = wb.active
            ws["C2"] = context_filters["name"]
            ws["C3"] = context_filters["lead_department"]
            ws["C4"] = context_filters["assigned_department"]
            ws["C5"] = context_filters["evaluation_result"]
            ws["C6"] = context_filters["report_statuses"]
            ws["C7"] = context_filters["period_label"]
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            vn_filename = "Báo cáo tổng hợp kết quả thực hiện chỉ tiêu theo đơn vị chủ trì.xlsx"
            ascii_filename = "Bao_cao_tong_hop_ket_qua_thuc_hien_chi_tieu_theo_don_vi_chu_tri.xlsx"
            encoded_vn_filename = quote(vn_filename)
            return HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": (
                        "attachment;"
                        f' filename="{ascii_filename}";'
                        f" filename*=UTF-8''{encoded_vn_filename}"
                    )
                },
            )

        # Số lượng dv thực hiện distinct union theo từng (dv chủ trì, tháng)
        # nên cần danh sách department_id cho từng (quota, tháng).
        report_dept_ids_by_key: dict[tuple[int, int], set[int]] = defaultdict(set)
        report_dept_rows = (
            QuotaReport.objects.filter(quota_id__in=quota_ids, period_id__in=period_ids)
            .values("quota_id", "period_id", "department_id")
            .distinct()
        )
        for r in report_dept_rows:
            report_dept_ids_by_key[(int(r["quota_id"]), int(r["period_id"]))].add(int(r["department_id"]))

        # Nếu filter theo `report_statuses`, chỉ tính các quota của tháng đó khi quota-tháng
        # có ít nhất 1 report thuộc các trạng thái được chọn.
        allowed_quota_period_keys: set[tuple[int, int]] | None = None
        if report_statuses:
            allowed_quota_period_keys = set(
                QuotaReport.objects.filter(
                    quota_id__in=quota_ids,
                    period_id__in=period_ids,
                    status__in=report_statuses,
                )
                .values_list("quota_id", "period_id")
                .distinct()
            )

        # Agg expected/actual theo (quota, tháng) để đánh giá từng chỉ tiêu.
        report_sums = (
            QuotaReport.objects.filter(quota_id__in=quota_ids, period_id__in=period_ids)
            .values("quota_id", "period_id")
            .annotate(
                expected_total=Coalesce(Sum("expected_value"), 0),
                actual_total=Coalesce(Sum("actual_value"), 0),
            )
        )

        # =====================
        # 2. TỔNG HỢP VÀO CẤU TRÚC DỮ LIỆU
        # =====================
        agg: dict[tuple[int, int], dict[str, Any]] = {}
        # key: (lead_department_id, period_id)

        for r in report_sums:
            quota_id = int(r["quota_id"])
            period_id = int(r["period_id"])

            if allowed_quota_period_keys is not None and (quota_id, period_id) not in allowed_quota_period_keys:
                continue

            lead_info = leader_by_quota.get(quota_id)
            if not lead_info:
                continue

            expected_total = int(r["expected_total"] or 0)
            actual_total = int(r["actual_total"] or 0)
            evaluation_result = expected_total > 0 and actual_total >= expected_total

            if evaluation_filter is not None and evaluation_filter != evaluation_result:
                continue

            group_key = (lead_info["lead_department_id"], period_id)
            if group_key not in agg:
                agg[group_key] = {
                    "total": 0,
                    "passed": 0,
                    "failed": 0,
                    "dept_set": set(),
                }

            agg[group_key]["total"] += 1
            if evaluation_result:
                agg[group_key]["passed"] += 1
            else:
                agg[group_key]["failed"] += 1

            # Union số lượng đơn vị thực hiện tối đa trong tháng cho dv chủ trì
            # của các chỉ tiêu được tính.
            agg[group_key]["dept_set"].update(report_dept_ids_by_key.get((quota_id, period_id), set()))

        # Chuyển agg -> rows để render
        # Sort theo tháng (desc như cách hiển thị filter) rồi sort dv chủ trì theo tên.
        rows: list[dict[str, Any]] = []
        for p in periods:
            # Lấy danh sách dv chủ trì có dữ liệu cho tháng p
            items = [
                (lead_id, period_id, data)
                for (lead_id, period_id), data in agg.items()
                if period_id == p.id
            ]
            if not items:
                continue

            # map lead_id -> name (lấy từ leader_by_quota)
            lead_name_by_id: dict[int, str] = {}
            for qid, info in leader_by_quota.items():
                lead_name_by_id.setdefault(info["lead_department_id"], info["lead_department_name"])

            items_sorted = sorted(
                items,
                key=lambda t: lead_name_by_id.get(t[0], ""),
            )
            for lead_id, period_id, data in items_sorted:
                rows.append(
                    {
                        "period_label": period_label_by_id.get(period_id, ""),
                        "lead_department_id": lead_id,
                        "lead_department_name": lead_name_by_id.get(lead_id, ""),
                        "total": int(data["total"]),
                        "passed": int(data["passed"]),
                        "failed": int(data["failed"]),
                        "max_units": len(data["dept_set"]),
                    }
                )

        # =====================
        # 3. RENDER EXCEL
        # =====================
        wb = load_workbook(self.TEMPLATE_PATH)
        ws = wb.active

        # Cột C chứa giá trị filter
        ws["C2"] = context_filters["name"]
        ws["C3"] = context_filters["lead_department"]
        ws["C4"] = context_filters["assigned_department"]
        ws["C5"] = context_filters["evaluation_result"]
        ws["C6"] = context_filters["report_statuses"]
        ws["C7"] = context_filters["period_label"]

        ws.column_dimensions["C"].width = 30

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_black = Side(style="thin", color="000000")
        table_border = Border(
            left=thin_black,
            right=thin_black,
            top=thin_black,
            bottom=thin_black,
        )

        start_row = 11
        total_rows = max(len(rows), 0)
        if len(rows) > 1:
            ws.insert_rows(start_row + 1, len(rows) - 1)

        for i, row in enumerate(rows, start=1):
            r = start_row + i - 1

            ws.cell(r, 1).value = i
            ws.cell(r, 2).value = row["period_label"]
            ws.cell(r, 3).value = row["lead_department_name"]
            ws.cell(r, 4).value = row["total"]
            ws.cell(r, 5).value = row["passed"]
            ws.cell(r, 6).value = row["failed"]
            ws.cell(r, 7).value = row["max_units"]

            for col in range(1, 8):  # A -> G
                cell = ws.cell(r, col)
                cell.border = table_border
                if col in (2, 3):
                    cell.alignment = left
                else:
                    cell.alignment = center

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        vn_filename = "Báo cáo tổng hợp kết quả thực hiện chỉ tiêu theo đơn vị chủ trì.xlsx"
        ascii_filename = "Bao_cao_tong_hop_ket_qua_thuc_hien_chi_tieu_theo_don_vi_chu_tri.xlsx"
        encoded_vn_filename = quote(vn_filename)

        return HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    "attachment;"
                    f' filename="{ascii_filename}";'
                    f" filename*=UTF-8''{encoded_vn_filename}"
                )
            },
        )

