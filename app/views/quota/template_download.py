from io import BytesIO
from pathlib import Path

from django.contrib.auth.decorators import permission_required
from django.http import Http404, HttpResponse
from django.utils.decorators import method_decorator
from django.views import View
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

from ...handlers import department

@method_decorator(permission_required("app.add_quota"), name="dispatch")
class QuotaTemplateDownloadView(View):
    def get(self, request, *args, **kwargs):
        template_path = (
            Path(__file__).resolve().parents[3]
            / "static"
            / "templates"
            / "bulk_quota.xlsx"
        )

        if not template_path.exists():
            raise Http404("Template file not found.")

        wb = load_workbook(template_path)

        # đổi tên sheet này theo file thật của bạn nếu khác
        guide_sheet_name = "Hướng dẫn"
        if guide_sheet_name not in wb.sheetnames:
            raise Http404(f"Template sheet '{guide_sheet_name}' not found.")

        ws = wb[guide_sheet_name]

        departments = department.get_all_departments()

        # ===== vùng bảng đơn vị bên phải =====
        # Theo ảnh bạn gửi:
        # N = STT, O = Tên đơn vị, P = Tên viết tắt, Q = Nhóm đơn vị
        header_row = 3
        start_row = 4
        start_col = 13  # cột M

        max_clear_row = max(ws.max_row, start_row + len(departments) + 50)

        # Xóa dữ liệu cũ vùng N4:Q...
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=max_clear_row,
            min_col=start_col,
            max_col=start_col + 3,
        ):
            for cell in row:
                cell.value = None
                cell.border = Border()
                cell.alignment = Alignment()

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(vertical="center", wrap_text=True)

        # Giữ format header N3:Q3
        for col in range(start_col, start_col + 4):
            cell = ws.cell(row=header_row, column=col)
            cell.border = border
            cell.alignment = center_alignment

        # Ghi dữ liệu đơn vị
        for idx, dept in enumerate(departments, start=1):
            row = start_row + idx - 1

            dept_name = getattr(dept, "name", "") or ""
            dept_short_name = getattr(dept, "short_name", "") or ""
            dept_group = getattr(dept, "type", "") or ""

            stt_cell = ws.cell(row=row, column=start_col, value=idx)           # N
            name_cell = ws.cell(row=row, column=start_col + 1, value=dept_name) # O
            short_cell = ws.cell(row=row, column=start_col + 2, value=dept_short_name) # P
            group_cell = ws.cell(row=row, column=start_col + 3, value=dept_group) # Q

            for cell in (stt_cell, name_cell, short_cell, group_cell):
                cell.border = border

            stt_cell.alignment = center_alignment
            name_cell.alignment = left_alignment
            short_cell.alignment = center_alignment
            group_cell.alignment = center_alignment

        # Set width đẹp hơn
        ws.column_dimensions["M"].width = 8
        ws.column_dimensions["N"].width = 40
        ws.column_dimensions["O"].width = 20
        ws.column_dimensions["P"].width = 16

        

        last_data_row = start_row + len(departments) - 1
        if last_data_row >= start_row:
            for row in range(start_row, last_data_row + 1):
                ws.row_dimensions[row].height = 30

        # Set height đẹp hơn
        ws.row_dimensions[2].height = 110  # dòng hướng dẫn dài
        ws.row_dimensions[3].height = 45   # header
        ws.row_dimensions[4].height = 45   # dòng ví dụ

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="mau_nhap_chi_tieu.xlsx"'
        return response