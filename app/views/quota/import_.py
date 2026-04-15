from datetime import date
from http import HTTPStatus
from io import BytesIO
from django.db import transaction
from django.db.models.fields import uuid
from django.contrib import messages
from django.shortcuts import render
from django.views import View
from django.contrib.auth.decorators import permission_required
from django.utils.timezone import datetime, timedelta
from django.utils.decorators import method_decorator
from django.utils.dateparse import parse_datetime
from openpyxl import load_workbook

from ...handlers import period, department
from ...models import Period, Quota, QuotaAssignment, QuotaReport
from utils.exception import ImportException

MAX_IMPORT = 100

LABEL_TYPE_MAP = {
    label: value
    for value, label in Quota.Type.choices
}

@method_decorator(permission_required("app.add_quota"), name="dispatch")
class QuotaImportView(View):
    template_name = "quota/import.html"

    def get(self, request, *args, **kwargs):
        return render(
            request,
            self.template_name,
            {
                'errors': {},
            },
        )

    def post(self, request, *args, **kwargs):
        errors = {}
        status = HTTPStatus.OK
        file = request.FILES.get('file', None)
        if not file:
            errors['file'] = 'Vui lòng tải lên file đính kèm'
            status = HTTPStatus.UNPROCESSABLE_ENTITY
        else:
            try:
                data, row_errors = self.validate_file(file)
                if row_errors:
                    status = HTTPStatus.UNPROCESSABLE_ENTITY
                    errors['rows'] = row_errors
                else:
                    with transaction.atomic():
                        quotas = []
                        quota_assignments = []
                        quota_reports = []
                        for row in data:
                            quota_uid = uuid.uuid4()
                            quota = Quota(
                                id=quota_uid,
                                name=row['name'],
                                type=row['type'],
                                register_guide=row['register_guide'],
                                submit_guide=row['submit_guide'],
                                target_percent=row['target_percent'] / 100,
                                issued_at=row['issued_at'],
                                expired_at=row['expired_at'],
                                department_id=row['lead_department_id'],
                            ).on_behalf_of(request.user)
                            quotas.append(quota)
                            for assigned_department_id in row['assigned_department_ids']:
                                assignment = QuotaAssignment(
                                    quota_id=quota_uid,
                                    department_id=assigned_department_id,
                                ).on_behalf_of(request.user)
                                quota_assignments.append(assignment)
                            period = Period.objects.filter(year=row['issued_at'].year, month=row['issued_at'].month).first()
                            if period:
                                for assigned_department_id in row['assigned_department_ids']:
                                    report = QuotaReport(
                                        quota_id=quota_uid,
                                        period=period,
                                        department_id=assigned_department_id,
                                        status=QuotaReport.Status.NOT_SENT,
                                    ).on_behalf_of(request.user)
                                    quota_reports.append(report)
                        Quota.objects.bulk_create(quotas)
                        QuotaAssignment.objects.bulk_create(quota_assignments)
                        QuotaReport.objects.bulk_create(quota_reports)
                        messages.success(request, 'Nhập chỉ tiêu thành công')
            except ImportException as e:
                errors['file'] = str(e)
                status = HTTPStatus.UNPROCESSABLE_ENTITY
            except Exception:
                messages.error(request, 'File không hợp lệ')
                status = HTTPStatus.INTERNAL_SERVER_ERROR
        return render(
            request,
            'quota/import_form.html',
            {
                "errors": errors,
            },
            status=status,
        )

    def validate_file(self, file):
        in_memory = file.read()
        wb = load_workbook(filename=BytesIO(in_memory))
        ws = wb['Data']
        max_row = ws.max_row
        if max_row <= 2:
            raise ImportException('File không có dữ liệu')
        if max_row > MAX_IMPORT:
            raise ImportException(f'Mỗi lần chỉ có thể tạo tối đa {MAX_IMPORT} chỉ tiêu')
        errors = []
        data = []
        deparment_map = {
            department.short_name: department.id for department in department.get_all_departments()
        }
        latest_period = period.get_latest_period()

        def to_str(v):
            if v is None:
                return ""
            if isinstance(v, str):
                return v.strip()
            return str(v)

        def validate_type(v):
            if not v:
                return None, "Cách tính không được để trống"
            if v not in LABEL_TYPE_MAP:
                return None, f"Cách tính '{v}' không hợp lệ"
            return LABEL_TYPE_MAP[v], None

        def validate_target_percent(v):
            if not v:
                return None, "Tỉ lệ được giao không được để trống"
            try:
                v = float(v)  # cho phép số thập phân
            except ValueError:
                return None, "Tỉ lệ được giao phải là số"

            if v <= 1 or v > 100:
                return None, "Tỉ lệ được giao phải > 1 và <= 100"

            return v, None

        def validate_issued_at(v):
            if not v:
                return None, "Ngày ban hành không được để trống"
            try:
                v = parse_datetime(v).date()
            except Exception:
                return None, "Ngày ban hành không hợp lệ"
            start_date = datetime(latest_period.year, latest_period.month, 1).date()
            if v < start_date:
                return None, f"Ngày ban hành không được nhỏ hơn {start_date.strftime('%d-%m-%Y')}"
            v.replace(day=1)
            return v, None

        def validate_expired_at(v):
            if not v:
                return None, "Ngày hết hiệu lực không được để trống"
            try:
                v = parse_datetime(v).date()
            except Exception as e:
                return None, "Ngày hết hiệu lực không hợp lệ"
            v = (datetime(v.year, v.month + 1, 1) - timedelta(days=1)).date()
            return v, None

        def validate_lead_department_short_name(v):
            if not v:
                return None, "Đơn vị chủ trì không được để trống"
            if v not in deparment_map:
                return None, "Không tìm thấy đơn vị với tên viết tắt: " + v
            return deparment_map[v], None

        def validate_assigned_department_short_names(v):
            if not v:
                return None, "Đơn vị thực hiện không được để trống"
            names = v.split(',')
            msg = "Không tìm thấy đơn vị với tên viết tắt: "
            error_names = []
            deparment_ids = []
            for name in names:
                name = name.strip()
                if not name:
                    continue
                if name not in deparment_map:
                    error_names.append(name)
                else:
                    deparment_ids.append(deparment_map[name])
            if error_names:
                msg += ", ".join(error_names)
                return None, msg
            return deparment_ids, None

        def validate_row_data(row):
            if row['lead_department_id'] in row['assigned_department_ids']:
                return 'lead_department_id', "Đơn vị chủ trì không thể đồng thời là đơn vị thực hiện"
            if row['issued_at'] > row['expired_at']:
                return 'expired_at', "Ngày hết hiệu lực không được nhỏ hơn ngày ban hành"
            return None, None

        columns = { 
            'name' : {
                'ordinal': 1,
                'label': 'Tên chỉ tiêu',
                'sample_value': 'Tỉ lệ chuyển đổi số',
                'validate': lambda v: (v, None) if v else (None, "Tên chỉ tiêu không được để trống"),
            },
            'type' : {
                'ordinal': 2,
                'label': 'Cách tính',
                'sample_value': 'Số liệu tại thời điểm báo cáo',
                'validate': validate_type,
            },
            'register_guide' : {
                'ordinal': 3,
                'label': 'Nội dung chỉ tiêu phải thực hiện',
                'sample_value': 'Số quy trình cần chuyển đổi số',
                'validate': lambda v: (v, None) if v else (None, "Nội dung chỉ tiêu phải thực hiện không được để trống"),
            },
            'submit_guide' : {
                'ordinal': 4,
                'label': 'Nội dung kết quả thực hiện',
                'sample_value': 'Số quy trình đã chuyển đổi số thành công',
                'validate': lambda v: (v, None) if v else (None, "Nội dung kết quả thực hiện không được để trống"),
            },
            'target_percent' : {
                'ordinal': 5,
                'label': 'Tỉ lệ được giao',
                'sample_value': 100,
                'validate': validate_target_percent,
            },
            'issued_at' : {
                'ordinal': 6,
                'label': 'Ngày ban hành',
                'sample_value': datetime(latest_period.year, latest_period.month, 1).date().strftime("%d-%m-%Y"),
                'validate': validate_issued_at,
            },
            'expired_at' : {
                'ordinal': 7,
                'label': 'Ngày hết hiệu lực',
                'sample_value': (datetime(latest_period.year, latest_period.month + 1, 1) - timedelta(days=1)).date().strftime("%d-%m-%Y"),
                'validate': validate_expired_at,
            },
            'lead_department_id' : {
                'ordinal': 8,
                'label': 'Đơn vị chủ trì',
                'sample_value': 'PV01',
                'validate': validate_lead_department_short_name,
            },
            'assigned_department_ids' : {
                'ordinal': 9,
                'label': 'Đơn vị thực hiện',
                'sample_value': "CAPHH, CAXLS_VT",
                'validate': validate_assigned_department_short_names,
            },
        }
        for r in range(2, max_row):
            row_data = {}
            all_empty = True
            row_error_count = 0
            for column, config in columns.items():
                raw = ws.cell(r, config['ordinal']).value or ''
                str_value = to_str(raw)
                if str_value:
                    all_empty = False
                validate = config.get('validate', lambda v: (True, None))
                value, error = validate(str_value)
                if error:
                    row_error_count += 1
                    errors.append({
                        'row_number': r,
                        'column_label': config['label'],
                        'input_value': raw.strftime("%d-%m-%Y") if isinstance(raw, date) else raw,
                        'sample_value': config['sample_value'],
                        'message': error,
                    })
                else:
                    row_data[column] = value
            if all_empty:
                if row_error_count:
                    errors = errors[:-row_error_count]
                break
            if row_error_count:
                continue
            err_column, error = validate_row_data(row_data)
            if error:
                errors.append({
                    'row_number': r,
                    'column_label': columns[err_column]['label'],
                    'input_value': row_data[err_column].strftime("%d-%m-%Y") if isinstance(row_data[err_column], date) else row_data[err_column],
                    'sample_value': columns[err_column]['sample_value'],
                    'message': error,
                })
            else:
                data.append(row_data)
        if errors:
            return None, errors
        if not data:
            raise ImportException('File không có dữ liệu')
        return data, None

