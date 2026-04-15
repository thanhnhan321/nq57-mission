from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from collections import defaultdict

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

from ...models import Period, Quota, QuotaReport
from .list import get_common_context


def _period_filter_label(request) -> str:
    """Nhãn kỳ từ query (period_ids); nếu trống thì tháng hiện tại nếu có trong DB."""
    raw = [v for v in request.GET.getlist("period_ids", []) if v]
    ids = [int(x) for x in raw if str(x).isdigit()]
    if not ids:
        today = timezone.localdate()
        p = Period.objects.filter(year=today.year, month=today.month).values_list("id", flat=True).first()
        if p is not None:
            ids = [p]
    if not ids:
        return "Tất cả"
    periods = Period.objects.filter(id__in=ids).order_by("-year", "-month")
    return ", ".join(f"Tháng {p.month:02d}/{p.year}" for p in periods)


def export_quota_excel(request):
    template_path = Path(__file__).resolve().parent / "summary_quota.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    # Cột C chứa cả nhãn filter (C2:C7) và tên chỉ tiêu (cột 3) nên cần đủ rộng
    # để hạn chế việc bị xuống dòng/che khuất.
    ws.column_dimensions["C"].width = 30

    # =====================
    # 1. LẤY DATA TỪ TABLE
    # =====================
    context = get_common_context(request)

    rows = context["rows"]  # đã transform sẵn

    filters = request.GET
    period_label = _period_filter_label(request)
    explicit_period_ids = [int(x) for x in request.GET.getlist("period_ids", []) if str(x).isdigit()]

    # =====================
    # 2. FILL FILTER (C2 → C7)
    # =====================
    ws["C2"] = filters.get("name", "Tất cả")
    ws["C3"] = filters.get("lead_department", "Tất cả")
    ws["C4"] = filters.get("assigned_department", "Tất cả")
    ws["C5"] = filters.get("evaluation_result", "Tất cả")
    ws["C6"] = ", ".join(filters.getlist("report_statuses")) or "Tất cả"
    ws["C7"] = period_label

    # =====================
    # 3. STYLE
    # =====================
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_black = Side(style="thin", color="000000")
    table_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)

    # =====================
    # 4. INSERT ROW
    # =====================
    start_row = 10
    total_rows = max(len(rows), 1)

    if total_rows > 1:
        ws.insert_rows(start_row + 1, total_rows - 1)

    # =====================
    # 5. WRITE DATA
    # =====================
    import re

    # Cột "Kỳ báo cáo" cần hiển thị theo từng chỉ tiêu:
    # - Nếu bạn đã chọn nhiều kỳ (period_ids), mỗi chỉ tiêu chỉ liệt kê các tháng trong selected_periods
    #   mà thực sự có bản ghi (QuotaReport) tương ứng.
    quota_ids = [row.get("id") for row in rows if row.get("id") is not None]
    period_labels_by_quota_id: dict[int, str] = {}

    if explicit_period_ids and quota_ids:
        ordered_periods = list(
            Period.objects.filter(id__in=explicit_period_ids).order_by("-year", "-month")
        )
        ordered_period_ids = [p.id for p in ordered_periods]
        period_label_by_id = {p.id: f"Tháng {p.month:02d}/{p.year}" for p in ordered_periods}

        # Giữ phạm vi dữ liệu giống như chỗ tổng hợp trong get_common_context:
        # - Nếu là leader cho quota đó: tính theo toàn bộ department_reports
        # - Nếu không: chỉ tính theo department của chính user.
        if request.user.is_superuser:
            quota_period_pairs = (
                QuotaReport.objects.filter(quota_id__in=quota_ids, period_id__in=explicit_period_ids)
                .values_list("quota_id", "period_id")
                .distinct()
            )
            periods_by_quota_id: dict[int, set[int]] = defaultdict(set)
            for qid, pid in quota_period_pairs:
                periods_by_quota_id[qid].add(pid)
        else:
            dept_id = request.user.profile.department_id
            leader_quota_ids = set(
                Quota.objects.filter(
                    id__in=quota_ids,
                    department_id=dept_id,
                ).values_list("quota_id", flat=True)
            )

            periods_by_quota_id = defaultdict(set)
            leader_quota_ids_list = [qid for qid in quota_ids if qid in leader_quota_ids]
            other_quota_ids_list = [qid for qid in quota_ids if qid not in leader_quota_ids]

            if leader_quota_ids_list:
                quota_period_pairs = (
                    QuotaReport.objects.filter(
                        quota_id__in=leader_quota_ids_list,
                        period_id__in=explicit_period_ids,
                    )
                    .values_list("quota_id", "period_id")
                    .distinct()
                )
                for qid, pid in quota_period_pairs:
                    periods_by_quota_id[qid].add(pid)

            if other_quota_ids_list:
                quota_period_pairs = (
                    QuotaReport.objects.filter(
                        quota_id__in=other_quota_ids_list,
                        department_id=dept_id,
                        period_id__in=explicit_period_ids,
                    )
                    .values_list("quota_id", "period_id")
                    .distinct()
                )
                for qid, pid in quota_period_pairs:
                    periods_by_quota_id[qid].add(pid)

        for qid in quota_ids:
            pids_for_quota = [pid for pid in ordered_period_ids if pid in periods_by_quota_id.get(qid, set())]
            label = ", ".join(period_label_by_id[pid] for pid in pids_for_quota)
            period_labels_by_quota_id[qid] = label if label else "-"

    for i, row in enumerate(rows, start=1):
        r = start_row + i - 1

        ws.cell(r, 1).value = i
        ws.cell(r, 2).value = period_labels_by_quota_id.get(row.get("id"), period_label)
        ws.cell(r, 3).value = row["name"]
        ws.cell(r, 4).value = row["lead_department_name"]
        ws.cell(r, 5).value = row.get("assigned_departments_full") or row["assigned_departments"]

        ws.cell(r, 6).value = f"{row['target_percent']*100:.2f}%"
        ws.cell(r, 7).value = row["total_expected"]
        ws.cell(r, 8).value = row["total_actual"]

        percent = row["completion_percent"] * 100
        ws.cell(r, 9).value = f"{percent:.2f}%"

        ws.cell(r, 10).value = "Đạt" if row["evaluation_result"] else "Không đạt"

        # ===== STATUS =====
        raw_status = row["report_statuses"]

        items = re.findall(r'>(.*?)<', raw_status)
        items = [i.strip() for i in items if i.strip()]

        formatted = "\n".join([f"- {item}" for item in items])

        ws.cell(r, 11).value = formatted

        # ===== ALIGNMENT =====
        for col in range(1, 12):
            cell = ws.cell(r, col)

            # Luôn có viền cho bảng export Excel
            cell.border = table_border

            if col in [2, 3, 4, 5, 11]:
                cell.alignment = left
            else:
                cell.alignment = center
    
    # =====================
    # 6. SAVE
    # =====================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    vn_filename = "Báo cáo tổng hợp kết quả thực hiện chỉ tiêu theo kỳ báo cáo.xlsx"
    # Một số trình duyệt/Windows không hiển thị đúng UTF-8 trong `filename`,
    # nên dùng `filename*` (RFC 5987) để đảm bảo tên có dấu.
    ascii_filename = "Bao_cao_tong_hop_ket_qua_thuc_hien_chi_tieu_theo_ky_bao_cao.xlsx"
    encoded_vn_filename = quote(vn_filename)

    return HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                "attachment;"
                f' filename="{ascii_filename}";'
                f" filename*=UTF-8''{encoded_vn_filename}"
            )
        }
    )