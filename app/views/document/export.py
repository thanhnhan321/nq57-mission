from datetime import date
from io import BytesIO
from pathlib import Path

from django.db.models import Q
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import permission_required
from django.views import View

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

import env

from ...models import Document, Department, Period


@method_decorator(permission_required("app.read_document"), name="dispatch")
class DocumentExportView(View):
    TEMPLATE_PATH = Path(__file__).resolve().parent / "template.xlsx"

    @staticmethod
    def resolve_department_name_from_filter_value(value: str) -> str:
        value = (value or "").strip()
        if not value:
            return ""

        dept = Department.objects.filter(id=value).values("name").first()
        if dept:
            return dept["name"]

        return value

    @staticmethod
    def parse_multi_values(request, key: str) -> list[str]:
        values = request.GET.getlist(key)
        if values:
            return [str(v).strip() for v in values if str(v).strip() not in ("", "all")]

        raw = request.GET.get(key)
        if not raw:
            return []

        if isinstance(raw, str) and "," in raw:
            return [part.strip() for part in raw.split(",") if part.strip() and part.strip() != "all"]

        raw = str(raw).strip()
        return [raw] if raw and raw != "all" else []

    @staticmethod
    def resolve_period_labels(period_ids: list[str]) -> str:
        if not period_ids:
            return "Tất cả"

        normalized_ids = []
        for value in period_ids:
            try:
                normalized_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        if not normalized_ids:
            return "Tất cả"

        periods = Period.objects.filter(id__in=normalized_ids).order_by("-year", "-month")
        labels = [f"{int(item.month):02d}/{int(item.year)}" for item in periods]
        return ", ".join(labels) if labels else "Tất cả"

    def get(self, request, *args, **kwargs):
        is_htmx = request.headers.get("HX-Request") == "true"
        if is_htmx:
            params = request.GET.urlencode()
            response = HttpResponse()
            response["HX-Redirect"] = env.HOST_URL + request.path + ("?" + params if params else "")
            return response

        period_ids = self.parse_multi_values(request, "period_ids")
        code = request.GET.get("code", "").strip()
        title = request.GET.get("title", "").strip()
        issued_by_raw = request.GET.get("issued_by", "").strip()
        status = request.GET.get("status", "").strip()

        issued_by_name = self.resolve_department_name_from_filter_value(issued_by_raw)
        period_label_text = self.resolve_period_labels(period_ids)

        queryset = Document.objects.select_related("type", "object", "period").all()

        if period_ids:
            queryset = queryset.filter(period_id__in=period_ids)

        if code:
            queryset = queryset.filter(code__icontains=code)

        if title:
            queryset = queryset.filter(title__icontains=title)

        if issued_by_name:
            queryset = queryset.filter(issued_by__iexact=issued_by_name)

        if status == "active":
            queryset = queryset.filter(Q(expired_at__isnull=True) | Q(expired_at__gte=date.today()))
        elif status == "expired":
            queryset = queryset.filter(expired_at__lt=date.today())

        queryset = queryset.order_by("-created_at")

        wb = load_workbook(str(self.TEMPLATE_PATH))
        ws = wb.active

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        header_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )

        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        column_widths = {
            "A": 6,
            "B": 16,
            "C": 15,
            "D": 20,
            "E": 15,
            "F": 30,
            "G": 15,
            "H": 18,
            "I": 18,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for col in range(1, 10):
            cell = ws.cell(row=8, column=col)
            cell.border = header_border
            cell.alignment = center_align

        # B2: Kỳ báo cáo
        # B3: Số văn bản
        # B4: Tên văn bản
        # B5: Đơn vị ban hành
        # B6: Tình trạng
        # C2:C6 là giá trị tương ứng
        ws["C2"] = period_label_text
        ws["C3"] = code or "Tất cả"
        ws["C4"] = title or "Tất cả"
        ws["C5"] = issued_by_name or "Tất cả"
        ws["C6"] = {
            "active": "Hiệu lực",
            "expired": "Hết hiệu lực",
        }.get(status, "Tất cả")

        start_row = 9

        for index, item in enumerate(queryset, start=1):
            row = start_row + index - 1

            ws[f"A{row}"] = index
            ws[f"B{row}"] = (
                f"{int(item.period.month):02d}/{int(item.period.year)}"
                if item.period
                else (item.issued_at.strftime("%m/%Y") if item.issued_at else "")
            )
            ws[f"C{row}"] = item.type.name if item.type else ""
            ws[f"D{row}"] = item.code or ""
            ws[f"E{row}"] = item.issued_at.strftime("%d/%m/%Y") if item.issued_at else ""
            ws[f"F{row}"] = item.issued_by or ""
            ws[f"G{row}"] = (
                "Hiệu lực"
                if (item.expired_at is None or item.expired_at >= date.today())
                else "Hết hiệu lực"
            )
            ws[f"H{row}"] = item.expired_at.strftime("%d/%m/%Y") if item.expired_at else ""
            ws[f"I{row}"] = item.created_at.strftime("%d/%m/%Y") if item.created_at else ""

            ws[f"A{row}"].alignment = center_align
            ws[f"B{row}"].alignment = center_align
            ws[f"E{row}"].alignment = center_align
            ws[f"G{row}"].alignment = center_align
            ws[f"H{row}"].alignment = center_align
            ws[f"I{row}"].alignment = center_align

            ws[f"C{row}"].alignment = left_align
            ws[f"D{row}"].alignment = left_align
            ws[f"F{row}"].alignment = left_align

            for col in range(1, 10):
                ws.cell(row=row, column=col).border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="danh_sach_van_ban.xlsx"'
        return response