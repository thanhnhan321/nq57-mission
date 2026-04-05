from io import BytesIO

from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required, permission_required
from django.views import View

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import env

from django.contrib.auth.models import User


@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('auth.view_user'), name='dispatch')
class UserExportView(View):

    def get(self, request, *args, **kwargs):
        is_htmx = request.headers.get("HX-Request") == "true"
        if is_htmx:
            params = request.GET.urlencode()
            response = HttpResponse()
            response["HX-Redirect"] = env.HOST_URL + request.path + ("?" + params if params else "")
            return response

        # Get filter parameters
        department_id = request.GET.get('department', '').strip()
        is_active = request.GET.get('is_active', '').strip()
        role = request.GET.get('role', '').strip()

        # Build queryset with filters
        queryset = User.objects.select_related('profile__department').all()

        if department_id and department_id.isdigit():
            queryset = queryset.filter(profile__department__id=department_id)

        if is_active == 'active':
            queryset = queryset.filter(is_active=True)
        elif is_active == 'inactive':
            queryset = queryset.filter(is_active=False)

        if role == 'admin':
            queryset = queryset.filter(is_superuser=True)
        elif role == 'member':
            queryset = queryset.filter(is_superuser=False)

        queryset = queryset.order_by('id')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Danh sách người dùng'

        # Define header style
        header_fill = PatternFill(start_color='F1F3F9', end_color='F1F3F9', fill_type='solid')
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define column widths
        column_widths = {
            'A': 6,    # STT
            'B': 15,   # Tên đăng nhập
            'C': 20,   # Họ tên
            'D': 20,   # Email
            'E': 15,   # Số điện thoại
            'F': 20,   # Đơn vị
            'G': 15,   # Trạng thái
            'H': 15,   # Vai trò
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Write header row
        headers = [
            'STT',
            'Tên đăng nhập',
            'Họ tên',
            'Email',
            'Số điện thoại',
            'Đơn vị',
            'Trạng thái',
            'Vai trò',
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for index, user in enumerate(queryset, start=1):
            profile = getattr(user, 'profile', None)
            full_name = (profile.full_name or '').strip() if profile else ''
            full_name = full_name or user.get_full_name() or user.username
            
            phone = (profile.phone or '').strip() if profile else ''
            phone = phone or '—'
            
            department = str(profile.department) if profile and profile.department else '—'
            
            is_active_text = 'Hoạt động' if user.is_active else 'Tạm khóa'
            role_text = 'Quản trị viên' if user.is_superuser else 'Thành viên'

            row = index + 1
            ws[f'A{row}'] = index
            ws[f'B{row}'] = user.username
            ws[f'C{row}'] = full_name
            ws[f'D{row}'] = user.email
            ws[f'E{row}'] = phone
            ws[f'F{row}'] = department
            ws[f'G{row}'] = is_active_text
            ws[f'H{row}'] = role_text

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return file
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="danh_sach_nguoi_dung.xlsx"'
        return response
